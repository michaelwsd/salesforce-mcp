import os
import base64
import requests as http_requests
from app import mcp, logger

TENANT_ID = os.environ.get("AZURE_TENANT_ID", "")
CLIENT_ID = os.environ.get("AZURE_CLIENT_ID", "")
CLIENT_SECRET = os.environ.get("AZURE_CLIENT_SECRET", "")
REFRESH_TOKEN = os.environ.get("ONEDRIVE_REFRESH_TOKEN", "")
ONEDRIVE_FOLDER = "GOWT Data Scrape"


def _get_onedrive_token() -> str:
    resp = http_requests.post(
        f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token",
        data={
            "client_id": CLIENT_ID,
            "client_secret": CLIENT_SECRET,
            "grant_type": "refresh_token",
            "refresh_token": REFRESH_TOKEN,
            "scope": "Files.ReadWrite offline_access",
        },
        timeout=30,
    )
    data = resp.json()
    token = data.get("access_token")
    if not token:
        raise RuntimeError(f"OneDrive token error: {data.get('error_description', data)}")
    return token


@mcp.tool()
def list_onedrive_files() -> list[dict]:
    """List all files in the GOWT Data Scrape folder on OneDrive.

    This folder contains the GOWT Excel spreadsheets:
    - GOWT_high.xlsx — one tab per GOWT High company with LinkedIn posts (monthly)
    - GOWT_mid_low.xlsx — combined news + FTE tracking for Medium/Low companies (quarterly)
    """
    token = _get_onedrive_token()
    resp = http_requests.get(
        f"https://graph.microsoft.com/v1.0/me/drive/root:/{ONEDRIVE_FOLDER}:/children",
        headers={"Authorization": f"Bearer {token}"},
        timeout=30,
    )
    if resp.status_code != 200:
        return {"error": f"Failed to list files ({resp.status_code}): {resp.text[:200]}"}
    items = resp.json().get("value", [])
    return [
        {
            "name": item["name"],
            "size": item.get("size"),
            "last_modified": item.get("lastModifiedDateTime"),
            "web_url": item.get("webUrl"),
        }
        for item in items
    ]


@mcp.tool()
def download_onedrive_file(filename: str) -> dict:
    """Download a file from the GOWT Data Scrape folder on OneDrive.

    Returns the file content as base64-encoded data. Use this to retrieve
    the GOWT Excel spreadsheets (GOWT_high.xlsx or GOWT_mid_low.xlsx).

    Args:
        filename: The filename to download (e.g. "GOWT_high.xlsx")
    """
    token = _get_onedrive_token()
    resp = http_requests.get(
        f"https://graph.microsoft.com/v1.0/me/drive/root:/{ONEDRIVE_FOLDER}/{filename}:/content",
        headers={"Authorization": f"Bearer {token}"},
        timeout=120,
    )
    if resp.status_code == 404:
        return {"error": f"{filename} not found in OneDrive/{ONEDRIVE_FOLDER}/"}
    if resp.status_code != 200:
        return {"error": f"Download failed ({resp.status_code}): {resp.text[:200]}"}

    logger.info(f"Downloaded {filename} ({len(resp.content)} bytes) from OneDrive")
    return {
        "filename": filename,
        "size": len(resp.content),
        "content_base64": base64.b64encode(resp.content).decode(),
    }


@mcp.tool()
def read_gowt_excel(filename: str, sheet_name: str | None = None, max_rows: int = 100) -> dict:
    """Read a GOWT Excel spreadsheet from OneDrive and return its contents as structured data.

    Downloads the file from OneDrive and parses it. Much more useful than
    download_onedrive_file because it returns the actual cell values.

    Available files:
    - "GOWT_high.xlsx" — tabs named after each GOWT High company, with LinkedIn posts
    - "GOWT_mid_low.xlsx" — FTE Tracking sheet + quarterly news sheets

    Args:
        filename: "GOWT_high.xlsx" or "GOWT_mid_low.xlsx"
        sheet_name: Specific sheet/tab to read. If omitted, returns all sheet names.
        max_rows: Maximum rows to return per sheet (default 100)
    """
    import io
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"error": "openpyxl not installed on server"}

    token = _get_onedrive_token()
    resp = http_requests.get(
        f"https://graph.microsoft.com/v1.0/me/drive/root:/{ONEDRIVE_FOLDER}/{filename}:/content",
        headers={"Authorization": f"Bearer {token}"},
        timeout=120,
    )
    if resp.status_code != 200:
        return {"error": f"Download failed ({resp.status_code}): {resp.text[:200]}"}

    wb = load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)

    if sheet_name is None:
        return {"filename": filename, "sheets": wb.sheetnames}

    if sheet_name not in wb.sheetnames:
        return {"error": f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"}

    ws = wb[sheet_name]
    rows = []
    headers = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c) if c is not None else f"col_{j}" for j, c in enumerate(row)]
            continue
        if i >= max_rows:
            break
        rows.append({headers[j]: c for j, c in enumerate(row) if j < len(headers)})

    wb.close()
    return {
        "filename": filename,
        "sheet": sheet_name,
        "headers": headers,
        "row_count": len(rows),
        "rows": rows,
    }
